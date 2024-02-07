from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from openpyxl import load_workbook
from openpyxl.styles import Font
import time
import os
import openpyxl

# Initialize Chrome browser
driver = webdriver.Chrome()

# Open Facebook login page
driver.get('https://www.facebook.com')

# Replace these with your Facebook credentials
username = 'rajkale743@gmail.com'
password = 'King@6369'

# Find the username and password fields and enter your credentials
username_field = driver.find_element(By.ID, 'email')
password_field = driver.find_element(By.ID, 'pass')

username_field.send_keys(username)
password_field.send_keys(password)

# Submit the login form
password_field.send_keys(Keys.RETURN)
time.sleep(5)

# Enter your search query here
user_query = input("Please enter your search query : ")   # for example you can search 'unsold car' as user_query

# Construct the URL based on the user's query (without specifying the country)
base_url = 'https://www.facebook.com/ads/library/?active_status=all&ad_type=all&country=ALL&sort_data[direction]=desc&sort_data[mode]=relevancy_monthly_grouped&media_type=all'
search_url = f'{base_url}&q={user_query}&sort_data[direction]=desc&sort_data[mode]=relevancy_monthly_grouped&search_type=keyword_unordered&media_type=all'

driver.get(search_url)

# File name
file_name = 'ad_data_3.xlsx'

# List to store the extracted URLs
multi_data_list = []
single_data_list = []

# Function to extract data and append it to the Excel file
def extract_and_append_data(selector, header):
    try:
        # Scroll down to the bottom of the page to load more content
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(10)  # Adjust the sleep time as needed
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
            time.sleep(5)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, selector))
        )

        elements = driver.find_elements(By.CLASS_NAME, selector)

        # Collecting data
        if elements:
            for element in elements:
                try:
                    data = element.get_attribute("href")    
                    if data:
                        multi_data_list.append(data)
                    # else:
                    #     multi_data_list.append("Not Found")
                except NoSuchElementException:
                    pass
        else:
            multi_data_list.append("Not Found")

        time.sleep(10)

        # delete useless items from multi_url list
        del multi_data_list[0]
        multi_data_list.remove('https://www.facebook.com/100008157758186')

        # Getting data into the excel
        if not multi_data_list:
            print(f"No {header} data found")
        else:
            # Check if the file exists, create it if not
            if not os.path.isfile(file_name):
                with open(file_name, 'w') as f:
                    f.write("")

            # Open the existing workbook or create a new one
            workbook = load_workbook(file_name) if os.path.getsize(file_name) > 0 else openpyxl.Workbook()

            # Select the desired sheet by name
            sheet = workbook.active

            # Find the next available column to append the new data
            next_column = 1
            while sheet.cell(row=1, column=next_column).value is not None:
                next_column += 1

            # Add a header for the new column if it doesn't exist
            header_cell = sheet.cell(row=1, column=next_column)
            if not header_cell.value:
                header_cell.value = header
                header_cell.font = Font(bold=True)  # Make the header bold

            # Append the collected data to the new column
            for row_index, data_item in enumerate(multi_data_list, start=2):
                sheet.cell(row=row_index, column=next_column, value=data_item)

            # Save the changes to the workbook
            workbook.save(file_name)
            
            print('multi_data_list : ', multi_data_list)
            print(f"\n{header} data successfully exported to '{file_name}'")

    except NoSuchElementException as e:
        print(f"Error: {e}")

def extract_and_append_data_by_link(header, *selector):
    # Store the main LinkedIn window handle
    main_window_handle = driver.current_window_handle

    for url in selector:
        try:
            driver.get(url)
            # Perform a cursor click before finding the 'ad_library' element
            action_chains = ActionChains(driver)
            action_chains.move_by_offset(0, 0).click().perform()
            time.sleep(10)
            
            # Locate and click the "About" link
            about_link = driver.find_element(By.LINK_TEXT, 'About')
            about_link.click()
            time.sleep(10)

            page_link = driver.find_element(By.LINK_TEXT, 'Page transparency')
            page_link.click()
            time.sleep(10)

            see_all_link = driver.find_element(By.XPATH, '/html/body/div[1]/div/div[1]/div/div[3]/div/div/div[1]/div[1]/div/div/div[4]/div/div/div/div[1]/div/div/div/div/div[2]/div/div/div/div/div[6]')
            see_all_link.click()
            time.sleep(10)

            library_link = driver.find_element(By.LINK_TEXT, 'Go to Ad Library')
            
            if about_link and page_link and see_all_link and library_link:
                library_link.click()

                # Switch to the single ad website tab
                for handle in driver.window_handles:
                    if handle != main_window_handle:
                        driver.switch_to.window(handle)
                        break

                # Get the current URL which should be the single ad page
                current_link = driver.current_url
                if current_link:
                    single_data_list.append(current_link)
                else:
                    single_data_list.append('NA') 
            else:
                single_data_list.append('NA')

            time.sleep(5)
            # Close the single ad page tab and switch back to the previous page
            driver.close()
            driver.switch_to.window(main_window_handle)

        except Exception as e:
            print(f"Error processing {url}: {str(e)}") 

    # Getting data into the excel
    if not single_data_list:
        print(f"No {header} data found")
    else:
        # Check if the file exists, create it if not
        if not os.path.isfile(file_name):
            with open(file_name, 'w') as f:
                f.write("")

        # Open the existing workbook or create a new one
        workbook = load_workbook(file_name) if os.path.getsize(file_name) > 0 else openpyxl.Workbook()

        # Select the desired sheet by name
        sheet = workbook.active

        # Find the next available column to append the new data
        next_column = 1
        while sheet.cell(row=1, column=next_column).value is not None:
            next_column += 1

        # Add a header for the new column if it doesn't exist
        header_cell = sheet.cell(row=1, column=next_column)
        if not header_cell.value:
            header_cell.value = header
            header_cell.font = Font(bold=True)  # Make the header bold

        # Append the collected data to the new column
        for row_index, data_item in enumerate(single_data_list, start=2):
            sheet.cell(row=row_index, column=next_column, value=data_item)

        # Save the changes to the workbook
        workbook.save(file_name)

        print('single_data_list : ', single_data_list)
        print(f"\n{header} data successfully exported to '{file_name}'")

my_class = 'xt0psk2.x1hl2dhg.xt0b8zv.x8t9es0.x1fvot60.xxio538.xjnfcd9.xq9mrsl.x1yc453h.x1h4wwuj.x1fcty0u'
column_name_1 = 'Multi Url'
column_name_2 = 'Single Url'

# Call the function for each data type
extract_and_append_data(my_class, column_name_1)
extract_and_append_data_by_link(column_name_2, *multi_data_list)

# Close the browser
driver.quit()