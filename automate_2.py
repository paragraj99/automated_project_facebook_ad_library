from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Initialize Chrome browser
driver = webdriver.Chrome()

# Enter your search query here
user_query = input("Please enter your search query : ")

# Construct the URL based on the user's query (without specifying the country)
base_url = 'https://www.facebook.com/ads/library/?active_status=all&ad_type=all&country=ALL&sort_data[direction]=desc&sort_data[mode]=relevancy_monthly_grouped&media_type=all'
search_url = f'{base_url}&q={user_query}&sort_data[direction]=desc&sort_data[mode]=relevancy_monthly_grouped&search_type=keyword_unordered&media_type=all'

driver.get(search_url)

# File name
file_name = 'ad_data_2.xlsx'

# Function to extract data and append it to the Excel file
def extract_and_append_data(selector, header):
    data_list = []
    
    try:
        # Scroll down to the bottom of the page to load more content
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(10)  # Adjust the sleep time as needed
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
            time.sleep(5)

        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, selector))
        )

        elements = driver.find_elements(By.CLASS_NAME, selector)

        # Collecting data
        if elements:
            for element in elements:
                try:
                    data = element.get_attribute("href")    
                    if data:
                        data_list.append(data)
                    # else:
                    #     data_list.append("Not Found")
                except NoSuchElementException:
                    pass
        else:
            data_list.append("Not Found")

        time.sleep(10)

        # delete useless items from multi_url list
        del data_list[0]
        data_list.remove('https://www.facebook.com/')


        # Getting data into the excel
        if not data_list:
            print(f"No {header} data found")
        else:
            # Check if the file exists, create it if not
            if not os.path.isfile(file_name):
                with open(file_name, 'w') as f:
                    f.write("")

            # Open the existing workbook or create a new one
            workbook = load_workbook(file_name) if os.path.getsize(file_name) > 0 else openpyxl.Workbook()

            # Select the desired sheet by name
            sheet = workbook.active

            # Find the next available column to append the new data
            next_column = 1
            while sheet.cell(row=1, column=next_column).value is not None:
                next_column += 1

            # Add a header for the new column if it doesn't exist
            header_cell = sheet.cell(row=1, column=next_column)
            if not header_cell.value:
                header_cell.value = header
                header_cell.font = Font(bold=True)  # Make the header bold

            # Append the collected data to the new column
            for row_index, data_item in enumerate(data_list, start=2):
                sheet.cell(row=row_index, column=next_column, value=data_item)

            # Save the changes to the workbook
            workbook.save(file_name)

            print(f"{header} data successfully exported to '{file_name}'")

    except NoSuchElementException as e:
        print(f"Error: {e}")


my_class = 'xt0psk2.x1hl2dhg.xt0b8zv.x8t9es0.x1fvot60.xxio538.xjnfcd9.xq9mrsl.x1yc453h.x1h4wwuj.x1fcty0u'
column_name = 'Multi Url'

# Call the function for each data type
extract_and_append_data(my_class, column_name)

# Close the browser
driver.quit()