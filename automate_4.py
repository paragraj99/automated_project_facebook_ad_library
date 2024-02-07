from selenium import webdriver
from selenium.common.exceptions import NoSuchElementException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
import pandas as pd
import time
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
import os

# Initialize Chrome browser
driver = webdriver.Chrome()

# Enter the link to navigate to the search result URL 
search_url = "https://www.facebook.com/ads/library/?active_status=all&ad_type=all&country=ALL&view_all_page_id=117084628161364&search_type=page&media_type=all"

driver.get(search_url)

# File name
file_name = 'ad_data_4.xlsx'

# Note - choice1='Css' its compulsory that value is always 'Css'
choice1='Css'
choice2='Class'

# Function to extract data and append it to the Excel file
def extract_and_append_data(selector, header, choice):
    data_list = []
    
    try:
        # Scroll down to the bottom of the page to load more content
        last_height = driver.execute_script("return document.body.scrollHeight")
        while True:
            driver.execute_script("window.scrollTo(0, document.body.scrollHeight);")
            time.sleep(10)  # Adjust the sleep time as needed
            new_height = driver.execute_script("return document.body.scrollHeight")
            if new_height == last_height:
                break
            last_height = new_height
            time.sleep(5)

        if choice == 'Css':
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, selector))
            )

            elements = driver.find_elements(By.CSS_SELECTOR, selector)

            # Collecting data
            if elements:
                for element in elements:
                    try:
                        data = element.text.strip()
                        if data:
                            data_list.append(data)
                    except NoSuchElementException:
                        pass
            else:
                data_list.append("Not Found")

            time.sleep(5)
        
        else:
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CLASS_NAME, selector))
            )

            elements = driver.find_elements(By.CLASS_NAME, selector)

            # Collecting data
            if elements:
                for element in elements:
                    try:
                        data = element.get_attribute("href")    
                        if data:
                            data_list.append(data)
                        # else:
                        #     data_list.append("Not Found")
                    except NoSuchElementException:
                        pass
            else:
                data_list.append("Not Found")

            time.sleep(5)

        # Getting data into the excel
        if not data_list:
            print(f"No {header} data found")
        else:
            # Check if the file exists, create it if not
            if not os.path.isfile(file_name):
                with open(file_name, 'w') as f:
                    f.write("")

            # Open the existing workbook or create a new one
            workbook = load_workbook(file_name) if os.path.getsize(file_name) > 0 else openpyxl.Workbook()

            # Select the desired sheet by name
            sheet = workbook.active

            # Find the next available column to append the new data
            next_column = 1
            while sheet.cell(row=1, column=next_column).value is not None:
                next_column += 1


            # Add a header for the new column if it doesn't exist
            header_cell = sheet.cell(row=1, column=next_column)
            if not header_cell.value:
                header_cell.value = header
                header_cell.font = Font(bold=True)  # Make the header bold

            # Append the collected data to the new column
            for row_index, data_item in enumerate(data_list, start=2):
                sheet.cell(row=row_index, column=next_column, value=data_item)

            # Save the changes to the workbook
            workbook.save(file_name)

            print(f"{header} data successfully exported to '{file_name}'")

    except NoSuchElementException as e:
        print(f"Error: {e}")

# Call the function for each data type
extract_and_append_data('.x67bb7w', 'Library ID', choice1)
extract_and_append_data('.x1i64zmx', 'Status', choice1)
extract_and_append_data('._2fyh', 'Brand', choice1)
extract_and_append_data('._7jyr', 'Describe', choice1)
extract_and_append_data('._8jh2', 'Ad Title', choice1)
extract_and_append_data('._8jh3', 'Title Describe', choice1)
extract_and_append_data('x1hl2dhg.x1lku1pv.x8t9es0.x1fvot60.xxio538.xjnfcd9.xq9mrsl.x1yc453h.x1h4wwuj.x1fcty0u.x1lliihq', 'Image URL', choice2)

# Close the browser
driver.quit()